import os
import getpass
from datetime import datetime
from openpyxl import Workbook, load_workbook
from Utilites.ExcelRead.ExcelReader import get_file_path

class TestReportWriter:
    def __init__(self, feature_name):
        self.username = getpass.getuser()
        self.feature_name = feature_name
        self.date_str = datetime.now().strftime("%Y-%m-%d")
        self.test_case_report_path = get_file_path("testCaseReportPath")
        self.folder_path = os.path.join(self.test_case_report_path, self.date_str)
        os.makedirs(self.folder_path, exist_ok=True)
        self.results = []

    def add_result(self, test_set_name, test_case_id, status, remarks, time_str):
        """
        Adds a test result to the report.
        Author:
            Prasad Kamble
        """
        self.results.append({
            "test_set_name": test_set_name,
            "test_case_id": test_case_id,
            "status": status,
            "remarks": remarks,
            "time": time_str
        })

    def write_report(self, build_version):
        """
        Writes the test report to an Excel file.
        Author:
            Prasad Kamble
        """
        file_name = f"{build_version}_{self.date_str}.xlsx"
        file_path = os.path.join(self.folder_path, file_name)
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestCaseReport"
            # Write header only if new file
            ws.append(["Test Set Name", "Test case ID", "Status", "Remarks", "Time"])
        # Append new results
        for result in self.results:
            ws.append([
                result["test_set_name"],
                result["test_case_id"],
                result["status"],
                result["remarks"],
                result["time"]
            ])
        wb.save(file_path)
        print(f"Test report written to: {file_path}")  

    def write_test_result(self, FEATURE_NAME, TEST_CASE_ID, BUILD_VERSION, status, remarks, logger_utils=None):
        """
        Utility function to add a test result and write the report.

        Args:
            FEATURE_NAME (str): The feature name.
            TEST_CASE_ID (str): The test case ID.
            BUILD_VERSION (str): The build version.
            status (str): Test status ("Pass"/"Fail").
            remarks (str): Remarks or error message.
            logger_utils: Optional logger utility for logging.
        """
        report_writer = TestReportWriter(FEATURE_NAME)
        report_writer.add_result(
            test_set_name=FEATURE_NAME,
            test_case_id=TEST_CASE_ID,
            status=status,
            remarks=remarks,
            time_str=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        report_writer.write_report(BUILD_VERSION)
        log_msg = (
            "========================================================\n"
            f"Test result written: {status}\n"
            "========================================================"
        )
        if logger_utils:
            logger_utils.log(log_msg)
        print(f"Test case status: {status}")